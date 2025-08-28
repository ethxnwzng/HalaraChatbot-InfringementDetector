import json

import openpyxl
import requests

from util.log_util import logger

auth = """
Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJmcCIsImV4cCI6MTY4MTI0MzU3MS42MjA1OTUyLCJpYXQiOjE2ODEyNDI5NzEuNjIwNTk1MiwiZGF0YSI6IntcImNyZWF0ZWRUaW1lXCI6IDE2ODEyNDEyNDAuODY2OTk0OSwgXCJwcm9maWxlSWRcIjogXCJGK2VLc09XeUlFUGthTytnVTNuU0ZjNm5TaXJURStMa3o4UTlBK0g4UWNtdHJCZHVyNXF5c0pWWThZNFJvSXFLNVUrM3JzMnJ4aGptNHBEMzBDY0lCdz09MTZhYzAzMzFhZDhhZDdmNmRhMTNhYWE2NTdhYWFhNTQyMGVmNTBiZjlkYmI5MGFkNTRhZmNkOTEyNzBmMWIxOFwiLCBcImFub255bW91c1wiOiB0cnVlLCBcInRyYWNlclwiOiBcIjZCV1FURTBTNkxcIiwgXCJzY29wZVwiOiBbXCJHVUVTVFwiXSwgXCJzaXRlSWRcIjogXCJmcC11c1wiLCBcImJyYW5kSWRcIjogXCJmcFwiLCBcInNpdGVHcm91cFwiOiBcImZwXCIsIFwiZGF0YUNlbnRlcklkXCI6IFwiVVMtTlZcIiwgXCJnZW9SZWdpb25cIjogXCJVUy1OVlwiLCBcImVkZ2VzY2FwZVwiOiB7XCJyZWdpb25Db2RlXCI6IFwiQ0FcIn0sIFwiY2FydElkXCI6IFwieEx4Z1RWaThvNjlEVFB0QVRsSm1zRjNoYStXV2JRSlplYTRDekNRRC82MFcyVFVNNE9GRmJDODVKc3FORHp5MDVVKzNyczJyeGhqbTRwRDMwQ2NJQnc9PTZmYTcxNjgxODFlNDM0YmQyYmRhNGJhZTQyYTQ0OTAwMWFmMGE3MDJhMmZhNTQ1YTA2ZTYwNzNkYTY2ZWM0M2FcIn0ifQ.q_oWG9yso1csHs0mVFBDE0D2nQPjVrJyFw8FGzLgWV4
""".strip()
cookie = """
ss-disable-dy=1; urbn_disable_dy=1; SSLB=1; SSSC_A15=513.G7220876143170314413.1|65792.2259787:69594.2339333:70402.2359044:72894.2403417:73154.2409049:73482.2415406:73695.2419358:73790.2421530:73861.2422889; urbn_currency=USD; siteId=fp-us; urbn_edgescape_site_id=fp-us; urbn_country=US; urbn_geo_region=US-NV; urbn_site_id=fp-us; urbn_language=en-US; urbn_clear=true; urbn_channel=web; urbn_inventory_pool=US_DIRECT; urbn_uuid=91dbc5ef-390a-4b3c-905d-4f16aadb18a0; urbn_tracer=6BWQTE0S6L; urbn_data_center_id=US-NV; AKA_A2=A; akacd_ss1=555-123-4567~rv=69~id=86d3db9881633002eeea48ccac1e7717; urbn_site_choice=fp-us->fp-us; urbn_device_info=web|other|desktop; split_tag_control=Conversant; FP_ATTR=other; pxcts=e2f4d941-d89e-11ed-8acb-4c505176626b; _pxvid=e2f4cc7b-d89e-11ed-8acb-4c505176626b; _gid=GA1.2.202064897.555-123-4567; _svsid=0e578cf5ef486444773e9659bd939196; __attentive_id=a77955bcac184a15823c943d587b7774; _attn_=eyJ1Ijoie1wiY29cIjoxNjgxMjQxMjQ2NjE0LFwidW9cIjoxNjgxMjQxMjQ2NjE0LFwibWFcIjoyMTkwMCxcImluXCI6ZmFsc2UsXCJ2YWxcIjpcImE3Nzk1NWJjYWMxODRhMTU4MjNjOTQzZDU4N2I3Nzc0XCJ9In0=; __attentive_cco=1681241246615; _gcl_au=1.1.973542789.555-123-4567; __attentive_ss_referrer=ORGANIC; _fbp=fb.1.1681241246866.337262335; __attentive_dv=1; _pin_unauth=dWlkPU0yRTBOekExWVRRdE0yRTBOeTAwWmpVeUxUZzJaVFF0TmpVeU0ySTFOV0l6WmpZMQ; _tt_enable_cookie=1; _ttp=aJkpSfVfzKSh56cKsLqPIovn7SA; BVBRANDID=59f4bb49-3039-4bee-aafd-aef8b1ad2a24; BVBRANDSID=d7008668-56fd-4464-b7c6-d007e44afc58; cebs=1; _ce.s=v~9c67f0d8a1f91a7494fad444d8775934734b9fe6~vpv~0; __qca=P0-967836569-1681241247623; _ce.clock_event=1; _ce.clock_data=64,174.139.126.207,1; _pxhd=OMGgyjMv3MxfTDR1TppMWeA-0OE5CTb59eGZumnQFmXb0-OqC5ovuBJbXEvQpsJM26oonppy504WrhB8F40W9w==:x5IknIZAfB5QLk/Yr2N6D51KRszio0PuSkFz92ixJwQdBy049S2cWzu9boxHz8XiGh4JRPvRhAUpk9CEFXrHPeZ44dVCTC4DcoHNM4uWQCE=; urbn_device_type=LARGE; urbn_privacy_restriction_region=["CA"]; _gac_UA-10477387-2=1.555-123-4567.Cj0KCQiAx6ugBhCcARIsAGNmMbikc9jpKVtbxG917VdE8X7nIsglLAOjxXrt3EXG1P9gjxvDxPScpQ8aAjdZEALw_wcB; _gcl_aw=GCL.555-123-4567.Cj0KCQiAx6ugBhCcARIsAGNmMbikc9jpKVtbxG917VdE8X7nIsglLAOjxXrt3EXG1P9gjxvDxPScpQ8aAjdZEALw_wcB; _gcl_dc=GCL.555-123-4567.Cj0KCQiAx6ugBhCcARIsAGNmMbikc9jpKVtbxG917VdE8X7nIsglLAOjxXrt3EXG1P9gjxvDxPScpQ8aAjdZEALw_wcB; urbn_uuid_session=b81cf0fe-0317-4631-b582-33c598e3ee28; SSRT_A15=WLs1ZAADAA; RT="z=1&dm=freepeople.com&si=h748r33bz0d&ss=lgcnml13&sl=0&tt=0"; _uetsid=e4f50b80d89e11ed8be14773e1cceb2e; _uetvid=e4f538f0d89e11ed8607a7442f25493f; _gat_tealium_0=1; _ga=GA1.1.215912893.555-123-4567; urbn_auth_payload={"authToken":"eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJmcCIsImV4cCI6MTY4MTI0MzU3MS42MjA1OTUyLCJpYXQiOjE2ODEyNDI5NzEuNjIwNTk1MiwiZGF0YSI6IntcImNyZWF0ZWRUaW1lXCI6IDE2ODEyNDEyNDAuODY2OTk0OSwgXCJwcm9maWxlSWRcIjogXCJGK2VLc09XeUlFUGthTytnVTNuU0ZjNm5TaXJURStMa3o4UTlBK0g4UWNtdHJCZHVyNXF5c0pWWThZNFJvSXFLNVUrM3JzMnJ4aGptNHBEMzBDY0lCdz09MTZhYzAzMzFhZDhhZDdmNmRhMTNhYWE2NTdhYWFhNTQyMGVmNTBiZjlkYmI5MGFkNTRhZmNkOTEyNzBmMWIxOFwiLCBcImFub255bW91c1wiOiB0cnVlLCBcInRyYWNlclwiOiBcIjZCV1FURTBTNkxcIiwgXCJzY29wZVwiOiBbXCJHVUVTVFwiXSwgXCJzaXRlSWRcIjogXCJmcC11c1wiLCBcImJyYW5kSWRcIjogXCJmcFwiLCBcInNpdGVHcm91cFwiOiBcImZwXCIsIFwiZGF0YUNlbnRlcklkXCI6IFwiVVMtTlZcIiwgXCJnZW9SZWdpb25cIjogXCJVUy1OVlwiLCBcImVkZ2VzY2FwZVwiOiB7XCJyZWdpb25Db2RlXCI6IFwiQ0FcIn0sIFwiY2FydElkXCI6IFwieEx4Z1RWaThvNjlEVFB0QVRsSm1zRjNoYStXV2JRSlplYTRDekNRRC82MFcyVFVNNE9GRmJDODVKc3FORHp5MDVVKzNyczJyeGhqbTRwRDMwQ2NJQnc9PTZmYTcxNjgxODFlNDM0YmQyYmRhNGJhZTQyYTQ0OTAwMWFmMGE3MDJhMmZhNTQ1YTA2ZTYwNzNkYTY2ZWM0M2FcIn0ifQ.q_oWG9yso1csHs0mVFBDE0D2nQPjVrJyFw8FGzLgWV4","reauthToken":"eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJmcCIsImV4cCI6MTY5Njc5NDk3MS42MjEwMjcsImlhdCI6MTY4MTI0Mjk3MS42MjEwMjcsImRhdGEiOiJ7XCJjcmVhdGVkVGltZVwiOiAxNjgxMjQyOTcxLjYyMTAxMTcsIFwic2NvcGVcIjogW1wiR1VFU1RcIl0sIFwidHJhY2VyXCI6IFwiNkJXUVRFMFM2TFwiLCBcInByb2ZpbGVJZFwiOiBcIkYrZUtzT1d5SUVQa2FPK2dVM25TRmM2blNpclRFK0xrejhROUErSDhRY210ckJkdXI1cXlzSlZZOFk0Um9JcUs1VSszcnMycnhoam00cEQzMENjSUJ3PT0xNmFjMDMzMWFkOGFkN2Y2ZGExM2FhYTY1N2FhYWE1NDIwZWY1MGJmOWRiYjkwYWQ1NGFmY2Q5MTI3MGYxYjE4XCJ9In0.Jj5ZdSkZJQM-2ESHZBNg1wDe212jgKplfkxd9EI9B2k","reauthExpiresIn":15552000,"expiresIn":600,"scope":"GUEST","tracer":"6BWQTE0S6L","dataCenterId":"US-NV","geoRegion":"US-NV","edgescape":{"regionCode":"CA","country":"US","city":"SANJOSE","zipCodes":"95101,95103,95106,95108-95113,95115-95136,95138-95139,95141,95148,95150-95161,95164,95170,95172-95173,95190-95194,95196"},"trueClientIp":"174.139.126.207","createdAt":1681242971625,"authExpiresTime":555-123-4567.625,"reauthExpiresTime":555-123-4567.625}; SSID_BE=CQArHR1-AAAAAACYtDVkrWDCKJi0NWQBAAAAAADEuAF3mLQ1ZADRnD4gAQMa8yQAmLQ1ZAEA3x8BAZ7qJACYtDVkAQCFIAEBafgkAJi0NWQBAL4cAQNZrCQAmLQ1ZAEAAAEBAUt7IgCYtDVkAQAKHwEDLtskAJi0NWQBAAITAQEE_yMAmLQ1ZAEAwh0BA1nCJACYtDVkAQDaDwEDBbIjAJi0NWQBAA; _derived_epik=dj0yJnU9SUEzMWlaMlRCRldFbWxyTTk4c2lnWGxlTWdCTXlIem4mbj1FeXI4N05PTkNZNHJOUzFNU0lVTzVBJm09NyZ0PUFBQUFBR1ExdTF3JnJtPTEmcnQ9QUFBQUFHT0cyQ0kmc3A9NQ; __attentive_pv=38; OptanonConsent=isGpcEnabled=0&datestamp=Wed+Apr+12+2023+03:56:13+GMT+0800+(China+Standard+Time)&version=6.38.0&isIABGlobal=false&hosts=&consentId=91dbc5ef-390a-4b3c-905d-4f16aadb18a0&interactionCount=0&landingPath=https://www.freepeople.com/shop/fp-one-aloha-printed-wide-leg-pants/?countryCode=us&quantity=1&ref=languageSelect&groups=C0001:1,C0002:1,C0003:1,C0004:1; utag_main=v_id:018771c97ef70034ddda35abfeaa05075007106d00b3b$_sn:1$_se:89$_ss:0$_st:1681244774805$ses_id:1681241243384;exp-session$_pn:13;exp-session$isLoggedIn:false;exp-session$dc_visit:1$dc_event:13;exp-session$dc_region:us-east-1;exp-session; _ga_PDBRPFW49G=GS1.1.555-123-4567.1.1.555-123-4567.57.0.0; cebsp_=13; RT="z=1&dm=www.freepeople.com&si=370ee74f-1b40-47d7-9d15-c297a51a1873&ss=lgcnml13&sl=2&tt=gi1&rl=1&bcn=//17de4c1c.akstat.io/"; akavpau_a15_freepeople_com_vp=555-123-4567~id=f7d642df3f9242a6f6cfa942f186e102
""".strip()


def get_product_reviews(handle):
    total, fetch = None, 0
    # proxy = {'http': 'http://154.90.4.218:2000', 'https': 'http://154.90.4.218:2000'}
    proxy = None
    review_meta = {'average_score': 0, 'recommend_count': 0, 'not_recommend_count': 0, 'total_review_count': None}
    reviews = []
    offset = 0
    limit = 100

    while total is None or fetch < total:
        url = 'https://www.freepeople.com/api/catalog/v0/fp-us/product/{}/reviews?projection-slug=reviews&offset={}&limit={}&filter='.format(handle, offset, limit)
        headers = {'authorization': auth, 'cookie': cookie, 'x-urbn-channel': 'web', 'x-urbn-country': 'US',
                   'x-urbn-currency': 'USD', 'x-urbn-language': 'en-US'}
        resp = requests.get(url, headers=headers, proxies=proxy)
        if resp and resp.ok:
            resp_data = resp.json()
            # print(json.dumps(resp_data))
            if review_meta['total_review_count'] is None:
                if 'product' in resp_data and 'reviewStatistics' in resp_data['product']:
                    review_meta_data = resp_data['product']['reviewStatistics']
                    review_meta['average_score'] = review_meta_data['averageOverallRating']
                    review_meta['recommend_count'] = review_meta_data['recommendedCount']
                    review_meta['not_recommend_count'] = review_meta_data['notRecommendedCount']
                    review_meta['total_review_count'] = review_meta_data['totalReviewCount']
                    total = int(review_meta['total_review_count'])
                if 'results' in resp_data and len(resp_data['results']) > 0:
                    for each in resp_data['results']:
                        review = {'id': each['id'], 'score': each['rating'], 'review_at': each['submissionTime'],
                                  'title': each['title'], 'text': each['reviewText'], 'nickname': each['userNickname'],
                                  'recommend': each['isRecommended'], 'location': each['userLocation']}
                        reviews.append(review)
                        fetch += 1
        else:
            logger.error('[freepeople] error: {}'.format(resp.text))
            break
        print('fetch: {} / {}'.format(fetch, total))
        offset += limit
        # break
    return review_meta, reviews


def export_to_excel(handle):
    o1, o2 = get_product_reviews(handle)
    print(o1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['id', 'score', 'time', 'title', 'text', 'nickname', 'recommend', 'location'])
    for e in o2:
        ws.append([e['id'], e['score'], e['review_at'], e['title'], e['text'], e['nickname'], e['recommend'],
                   e['location']])
    wb.save('./data/freepeople_{}.xlsx'.format(h))


def construct_reviews(handle):
    wb = openpyxl.load_workbook('./freepeople_{}.xlsx'.format(handle))
    ws = wb.active
    for i in range(2, ws.max_row+1):
        score = int(ws.cell(i, 2).value)
        title = ws.cell(i, 4).value
        text = ws.cell(i, 5).value
        recommend = str(ws.cell(i, 7).value)
        # print(recommend)
        if recommend == 'True' or score >= 4:
            recommend = 'yes'
        elif recommend == 'False' or score <= 2:
            recommend = 'no'
        else:
            recommend = 'unknown'

        print('{}\n[title]: {}\n[content]: {}\n[recommend]: {}\n'.format(i-1, title, text, recommend))


if __name__ == '__main__':
    # h = 'goddess-lounge-pants'
    # h = 'chica-lyrical-flow-pant'
    h = 'fp-one-aloha-printed-wide-leg-pants'
    # h = 'wake-up-pants2'
    # h = 'lyla-linen-trousers'
    # h = 'fairy-chiffon-pants'
    # h = 'clover-wide-leg-solid-pants'
    # h = 'leo-pants'
    export_to_excel(h)
    # construct_reviews(h)


