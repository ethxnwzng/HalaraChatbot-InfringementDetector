import os
import re

import django
import openpyxl

from home.gpt.review_reader.enums import ExtractType
from util import time_util

os.environ['DJANGO_SETTINGS_MODULE'] = 'lark.settings'
django.setup()

from home.models import ObserverArchive


def export_review_dimensions():
    archives = ObserverArchive.objects.all()
    wb = openpyxl.Workbook()
    for archive in archives:
        archive_name = archive.archive_name
        wb.create_sheet(archive_name)
        ws = wb[archive_name]
        ws.append(['id', 'content', 'pros', 'cons', 'scenario'])
        reviews = archive.origin_info.get('reviews', [])
        process_info = archive.process_info
        map_pros = tackle_process_info(process_info, ExtractType.PROS)
        map_cons = tackle_process_info(process_info, ExtractType.CONS)
        map_scenario = tackle_process_info(process_info, ExtractType.SCENARIO)
        # print(map_pros, map_cons, map_scenario)
        for review in reviews:
            review_id = str(review['id'])
            pros_list, cons_list, scenario_list = map_pros.get(review_id, []), map_cons.get(review_id, []), map_scenario.get(review_id, [])
            ws.append([review_id, review['content'], '\n'.join(pros_list), '\n'.join(cons_list), '\n'.join(scenario_list)])

    wb.save('./reviews_points_{}.xlsx'.format(time_util.get_cnt_now().strftime(time_util.DATE_FORMAT_COMPACT)))


def tackle_process_info(process_info: dict, extract_type: ExtractType):
    map_result = {}
    key = 'batch_points_{}'.format(extract_type.name())
    if key not in process_info:
        return map_result
    for each in process_info[key]:
        for e in each:
            ids = e['ids']
            point = e['point']
            for id_ in ids:
                id_ = str(id_)
                if id_ not in map_result:
                    map_result[id_] = []
                map_result[id_].append(point)
    return map_result


def re_merge():
    wb = openpyxl.load_workbook('../data/tmp.xlsx')
    ws = wb.active
    pt = re.compile('^(\d+):.*$')
    set_dedup = set()
    for i in range(2, ws.max_row+1):
        point = ws.cell(i, 1).value
        detail = ws.cell(i, 3).value
        dedup = '{}::{}'.format(point, detail)
        if dedup in set_dedup:
            continue
        else:
            set_dedup.add(dedup)
        ids = []
        for line in detail.split('\n'):
            line = str(line).strip()
            if mt := pt.match(line):
               ids.append(mt.group(1))
        print('- {} ({})'.format(point, ', '.join(ids)))
        if len(set_dedup) >= 20:
            break


if __name__ == '__main__':
    # export_review_dimensions()
    re_merge()

