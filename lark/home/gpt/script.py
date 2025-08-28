import re
import os

import django

from lark.settings import BASE_DIR
from util.lark_util import Lark

os.environ['DJANGO_SETTINGS_MODULE'] = 'lark.settings'
django.setup()
import openpyxl

from home.config import constant
from home.gpt import dto
from home.gpt.review_reader import warehouse
from home.models import LarkUserWorkbench
from util.log_util import logger

pt = re.compile(r'^(.*)\(([^()]*)\).*$')
pt_colon = re.compile(r'^(.*): *([\d, ]+).*$')
pt_num_range = re.compile(r'^(\d+) *- *(\d+)$')
pt_filter_quote = re.compile(r'^.*filter_quote: (\d+)\D*$')
pt_num_has = re.compile(r'^.*\d+.*$')
pt_num_trim = re.compile(r'^\D*([\d\- ]+)\D*$')


def parse_analyze(raw):
    points, filter_quote = [], 0
    lines = raw.strip().split('\n')
    for line in lines:
        line = line.strip()
        # get filter_quote
        if 'Reviews Report' in line:
            if mt := pt_filter_quote.match(line):
                filter_quote = int(mt.group(1))
        if not line.startswith('-') or ('(' not in line and ':' not in line):
            continue
        line = line.strip(',.- ')
        if '(' in line and not line.endswith(')'):
            line += ')'
        # print(line)
        if mt := pt.match(line):
            point = str(mt.group(1)).strip('. ')
            ids_str = mt.group(2).strip().strip('[]').split(',')
        elif mt := pt_colon.match(line):
            point = str(mt.group(1)).strip('. ')
            ids_str = mt.group(2).strip().strip('[]').split(',')
        else:
            point, ids_str = None, None
            Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatgpt reviews error] parse analyze empty', raw)
        if point and ids_str:
            ids = []
            ids_raw = []
            for each in ids_str:
                each = each.strip().strip("'")
                if pt_num_has.match(each):
                    if mt_num_trim := pt_num_trim.match(each):
                        ids_raw.append(mt_num_trim.group(1))
            for each in ids_raw:
                ids += parse_num(each)
            ids = list(set(ids))
            if len(ids) <= filter_quote:
                continue
            ids.sort(key=lambda x: int(x))
            points.append({'point': point, 'ids': ids})
    # rank
    points.sort(key=lambda x: len(x['ids']), reverse=True)

    # for each in points:
    #     print(each['point'], each['ids'])

    return points


def read_analyze(dimension):
    if dimension not in warehouse.analyze:
        return None
    raw = warehouse.analyze[dimension]
    # parse
    return parse_analyze(raw)


def read_reviews(file_path):
    map_review = {}
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]
    for i in range(2, ws.max_row+1):
        id_ = ws.cell(i, 1).value
        if not id_:
            continue
        content = ws.cell(i, 2).value
        if not content:
            content = ''
        # print(str(id_), str(content))
        map_review[str(id_)] = str(content)
    return map_review


def export(file_path=None):
    dimensions = ['cons', 'scenario', 'pros', 'common']
    need_write_reviews, record_count = False, 0
    if file_path:
        map_review = read_reviews(file_path)
        wb = openpyxl.load_workbook(file_path)
        record_count = len(map_review)
    else:
        need_write_reviews = True
        record = LarkUserWorkbench.objects.filter(lark_user_id=constant.LARK_USER_ID_LUKE, work_symbol=dto.WORK_SYMBOL).first()
        excel_data = record.detail.get(dto.JSON_KEY_EXCEL_DATA, [])
        excel_name = record.detail.get(dto.JSON_KEY_EXCEL_NAME, 'tmp.xlsx')
        map_review = dto.get_db_map_reviews(excel_data)
        wb = openpyxl.Workbook()
        file_path = str(BASE_DIR) + '/home/gpt/data/{}'.format(excel_name)
        record_count = len(excel_data)
    hit_ids = set()
    for dimension in dimensions:
        logger.info('[ChatGPT Review] export dimension: {}'.format(dimension))
        point_list = read_analyze(dimension)
        if not point_list:
            continue
        if dimension in wb.sheetnames:
            wb.remove(wb[dimension])
        wb.create_sheet(dimension)
        ws = wb[dimension]
        ws.append(['', record_count, ''])
        ws.append(['point', 'count', 'ids', 'detail'])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 100
        for each in point_list:
            ids = list(set(each['ids']))
            ids.sort(key=lambda x: int(x))
            detail = ''
            for index in ids:
                hit_ids.add(index)
                detail += '{}: {}\n'.format(index, map_review.get(index, ''))
            count = len(ids)
            ws.append([each['point'], count, ', '.join(ids), detail.strip()])

    # write review detail
    if need_write_reviews:
        ws_detail = wb.worksheets[0]
        ws_detail.append(['id', 'content'])
        for id_, content in map_review.items():
            ws_detail.append([id_, content])

    # other
    other_reviews = []
    if 'other' in wb.sheetnames:
        wb.remove(wb['other'])
    wb.create_sheet('other')
    ws = wb['other']
    ws.column_dimensions['B'].width = 100
    ws.append(['id', 'review'])
    for k, v in map_review.items():
        if k not in hit_ids:
            other_reviews.append({'id': k, 'review': v})
    other_reviews.sort(key=lambda x: len(x['review']), reverse=True)
    for each in other_reviews:
        ws.append([each['id'], each['review']])
    wb.save(file_path)


def parse_num(raw):
    nums = []
    if mt := pt_num_range.match(raw):
        int_from = int(mt.group(1))
        int_to = int(mt.group(2))
        for i in range(int_from, int_to+1):
            nums.append(str(i))
    else:
        nums.append(str(raw))
    return nums


# 将中间评论trim掉id list节约tokens
def transform_analyze(raw, item_index, filter_quote=0):
    points_raw = parse_analyze(raw)
    # filter points which less quote
    if filter_quote:
        points = []
        for each in points_raw:
            if len(each['ids']) > filter_quote:
                points.append(each)
    else:
        points = points_raw
    text, map_link = '', {}
    for each in points:
        item_index += 1
        text += '{}. {}\n'.format(item_index, each['point'])
        map_link[item_index] = each['ids']
    return text.strip(), map_link


# 将结果从point ids切换到完整评论ids展示
def transform_result(reply, map_reviews):
    wrap_reply, reply_list = '', []
    points = parse_analyze(reply)
    for each in points:
        review_ids = []
        for id_ in each['ids']:
            review_ids += map_reviews.get(int(id_), [])
        review_ids = list(set(review_ids))
        review_ids.sort(key=lambda x: int(x))
        reply_list.append({'point': each['point'], 'ids': review_ids})
    reply_list.sort(key=lambda x: len(x['ids']), reverse=True)
    for each in reply_list:
        wrap_reply += '- {} ({})\n'.format(each['point'], ', '.join(each['ids']))
    return wrap_reply


def manual_merge(replies, merged=None):
    print(len(replies))
    index, map_ = 0, {}
    for reply in replies:
        t_, m_ = transform_analyze(reply, index)
        index += len(m_)
        map_.update(m_)
        print(t_)
    # print(map_)
    if merged:
        merged_ = warehouse.merged
        p_ = parse_analyze(merged_)
        # print(len(p_), p_)
        r_ = transform_result(merged_, map_)
        print(r_)


PT_REPLY_MULTI = re.compile(r'^.*\d+\), .*$')


def check_reply_ok(reply):
    if not reply:
        return True
    for line in reply.split('\n'):
        line = line.strip()
        if PT_REPLY_MULTI.match(line):
            return False
    return True


if __name__ == '__main__':
    # file_ = str(BASE_DIR) + '/home/gpt/data/广告评论_22Q14LG101_3449.xlsx'
    export(file_path=None)

    # print(o)
    # r = warehouse.replies
    # m = warehouse.merged
    # m = None
    # manual_merge(replies=r, merged=m)



