"""
NOTE: This file has been sanitized for portfolio display.
Replace placeholder values with actual credentials for production use.
"""

import copy
import random
import time
import traceback

import openpyxl
import tiktoken

from home import redis_key
from home.config import constant
from home.gpt import chat_client, meta, script, dto, algo_client
from home.gpt.review_reader import self_tag
from home.gpt.review_reader.enums import ArchiveType, ExtractType
from home.models import LarkUserWorkbench, ObserverArchive
from util import redis_util
from util.lark_util import Lark
from util.log_util import logger
from home.idea_bot import client as bot_client
from home.gpt.review_reader import meta as review_meta


# default params
DEFAULT_EXCEL_COLUMNS = '#1, #2'
DEFAULT_REVIEW_FORMAT = """
[id] #1
[content] #2
""".strip()

DEFAULT_OBJ_DESC = "women's clothing customer reviews"
DEFAULT_TIPS = "format: id / content"
PROMPT_CONS = "Cons(product defects and pain points with negative sentiment)"
PROMPT_SCENARIO = "Dressing-Scenario(usage scenarios and activities for which customers wear the clothing; if more than one is mentioned in a review, treat each as a separate)"
PROMPT_PROS = "Pros(product benefits and selling points with positive sentiment)"
PROMPT_COMMON = "Main Ideas"

DEFAULT_ASK_PROMPT = """
Below is the list of #object_desc# (#parentheses_tips#).
Please summarize the points of #extra_desc# in specific expression.
Requirements:
If more than one aspect is mentioned in a review, split to several points.
Each point of one aspect in a new line begins with dash symbol.
Mention the id of reviews that behind each point in parentheses after the point.
Just list the points and mentions without explanation or example.
If no related points just leave it empty.
Before giving the answer, check if there are similar points(aspects) can be merged to one, then do it.
""".strip()
DEFAULT_MERGE_PROMPT = """
Below is a list of points summarized from #object_desc# about #extra_desc#. Each line in the list is an item in the format:
<ID>. <Point>

Merge the points that express the same meaning or belong to the same categories specified in parentheses(shipping, sizing, price, return-policy, customer-service).
Requirements:
Don't miss any points.
Material or fabric issues should be split into specific points.
Each point in a new line.
Never merge to overly broad summaries like "other issues".

Then format the result as a list of items. Each item is in the format:
- <Point> (<list of related ID>)

Points list:
""".strip()
DEFAULT_SYS_PROMPT = """
""".strip()
DEFAULT_AUTO_MERGE = True
DEFAULT_TEMPERATURE = 0
DEFAULT_WORDS_BATCH = 3000

MAX_REVIEWS_BATCH = 100
MAX_WORDS_BATCH = 3000
MAX_TOKENS_IN_MERGE = 1500

USE_MODEL = meta.MODEL_REVIEW
encoding = tiktoken.encoding_for_model(USE_MODEL)


class RequestRunReview:
    def __init__(self):
        self.user_id = None
        self.review_format = ''
        self.ask_prompt = ''
        self.merge_prompt = ''
        self.system_prompt = ''
        self.temperature = 0
        self.auto_merge = False
        self.use_lark = False
        # max parts
        self.max_segments = None
        self.words_batch = MAX_WORDS_BATCH
        self.archive_name = ''
        self.archive_data = None
        self.extract_type = ExtractType.UNKNOWN


class RequestMergeReview:
    def __init__(self):
        self.user_id = None
        self.replies = []
        self.merge_prompt = ''
        self.system_prompt = ''
        self.temperature = 0
        self.use_lark = False
        self.map_reviews = {}
        self.archive_name = ''
        self.extract_type = ExtractType.UNKNOWN


def tackle_excel(user_id, file_obj, columns_str, max_row=3000, archive_type=ArchiveType.UNKNOWN):
    reviews = []
    # 1. parse file
    if not file_obj or not columns_str or not user_id:
        return True, reviews
    file_name = str(file_obj)
    columns = []
    for co in str(columns_str).strip().split(constant.COMMA):
        columns.append(co.strip())
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    count = 0
    for i in range(2, ws.max_row+1):
        data = {}
        for col in columns:
            index = int(col.strip('#'))
            each = ws.cell(i, index).value
            if each is None:
                continue
            data[col] = str(each)
        count += 1
        if count > max_row:
            break
        reviews.append(data)

    # 2. write in workflow db
    record = LarkUserWorkbench.objects.filter(lark_user_id=user_id, work_symbol=dto.WORK_SYMBOL).first()
    if not record:
        record = LarkUserWorkbench(lark_user_id=user_id, work_symbol=dto.WORK_SYMBOL, detail={})
    record.detail[dto.JSON_KEY_EXCEL_NAME] = file_name
    record.detail[dto.JSON_KEY_EXCEL_DATA] = reviews
    record.detail[dto.JSON_KEY_EXCEL_COLUMNS] = columns_str
    record.save()

    # 3. write in archive db
    if archive_type == ArchiveType.UNKNOWN:
        archive_type = get_archive_type_by_file_name(file_name)
    key_id = '#{}'.format(review_meta.INDEX_ID)
    key_content = '#{}'.format(review_meta.INDEX_CONTENT)
    origin_info = {'reviews': list(map(lambda x: {'id': x[key_id], 'content': x[key_content]}, reviews))}
    archive = ObserverArchive.objects.filter(archive_type=archive_type.value, archive_name=file_name, record_count=len(reviews)).first()
    if not archive:
        archive = ObserverArchive(archive_type=archive_type.value, archive_name=file_name, record_count=len(reviews),
                                  origin_url='', origin_info=origin_info, process_info={})
        archive.save()

    return True, reviews


def get_archive_type_by_file_name(file_name):
    if '商城评论' in file_name:
        archive_type = ArchiveType.MALL_REVIEWS
    elif '广告评论' in file_name:
        archive_type = ArchiveType.AD_REVIEWS
    else:
        archive_type = ArchiveType.UNKNOWN
    return archive_type


def write_process_info(req, replies=None, reports=None):
    archive_name = req.archive_name
    archive_type = get_archive_type_by_file_name(archive_name)
    extract_type = req.extract_type
    archive = ObserverArchive.objects.filter(archive_type=archive_type.value, archive_name=archive_name,
                                             record_count=len(req.archive_data)).first()
    if not archive:
        return False
    new_process_info = {}

    if replies:
        batch_points = []
        for reply in replies:
            points = script.parse_analyze(reply)
            batch_points.append(points)
        new_process_info.update({'replies_{}'.format(extract_type.name()): replies,
                                 'batch_points_{}'.format(extract_type.name()): batch_points})

    if reports:
        new_process_info.update({'reports_{}'.format(extract_type.name()): reports})

    if not archive.process_info:
        archive.process_info = {}
    archive.process_info.update(new_process_info)
    archive.save()
    return True


def run_review(req: RequestRunReview):
    replies, ok = [], True
    try:
        if not req.user_id:
            return None
        # 1. get user's review content
        reviews = []
        record = LarkUserWorkbench.objects.filter(lark_user_id=req.user_id, work_symbol=dto.WORK_SYMBOL).first()
        if not record:
            return None
        excel_data = record.detail.get(dto.JSON_KEY_EXCEL_DATA, [])
        map_reviews = dto.get_db_map_reviews(excel_data)
        excel_name = record.detail.get(dto.JSON_KEY_EXCEL_NAME, '')
        req.archive_name = excel_name
        req.archive_data = excel_data
        sys_words_count = len(req.ask_prompt.split(' ')) + len(req.system_prompt.split(' '))
        review_index, words_count, separate_index_list = 0, sys_words_count, [0]
        for each in excel_data:
            review_index += 1
            each_text = req.review_format
            for k, v in each.items():
                each_text = each_text.replace(k, str(v).strip())
            words_count += len(encoding.encode(each_text))
            if words_count > req.words_batch or (review_index - separate_index_list[-1]) > MAX_REVIEWS_BATCH:
                if review_index-1 == separate_index_list[-1]:
                    # 单个评论已经超过了，报错
                    logger.error('[chatGPT review run] one review exceed limit, user: {}, review: {} \n{}'.format(req.user_id, review_index, each_text))
                    Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT review run] one review exceed limit', 'user: {}, review: {} \n{}'.format(req.user_id, review_index, each_text))
                    continue
                # 截断分解提问
                separate_index_list.append(review_index-1)
                words_count = sys_words_count
            reviews.append(each_text)
        # 2. save params config
        record.detail[dto.JSON_KEY_REVIEW_FORMAT] = req.review_format
        record.detail[dto.JSON_KEY_ASK_PROMPT] = req.ask_prompt
        record.detail[dto.JSON_KEY_MERGE_PROMPT] = req.merge_prompt
        record.detail[dto.JSON_KEY_SYS_PROMPT] = req.system_prompt
        record.detail[dto.JSON_KEY_AUTO_MERGE] = req.auto_merge
        record.detail[dto.JSON_KEY_TEMPERATURE] = req.temperature
        record.detail[dto.JSON_KEY_WORDS_BATCH] = req.words_batch
        record.save()

        # 3. wrap full ask text
        user_prompt_list = []
        for i in range(len(separate_index_list)):
            index_from = separate_index_list[i]
            if i == len(separate_index_list) - 1:
                index_to = len(reviews)
            else:
                index_to = separate_index_list[i+1]
            user_prompt = '{}\n\n{}'.format(req.ask_prompt, '\n\n'.join(reviews[index_from:index_to]))
            body = {'index_from': index_from, 'index_to': index_to, 'content': user_prompt}
            user_prompt_list.append(body)
        if req.max_segments and len(user_prompt_list) > req.max_segments:
            replies.append('too many reviews to tackle in page, pls click use-lark-notify mode')
            return replies

        # 4. ask chatGPT
        for i in range(len(user_prompt_list)):
            body = user_prompt_list[i]
            prompt_list = ['user::{}'.format(body['content'].strip())]
            time.sleep(0.7 + random.randint(1, 5))
            retry, retry_max, hit, reply = 0, 3, False, ''
            while retry < retry_max:
                retry += 1
                hit, reply = chat_client.ask(prompt_list, system_prompt=req.system_prompt, model=USE_MODEL,
                                             temperature=req.temperature, auto_retry=True)
                if hit and script.check_reply_ok(reply):
                    break
            if not hit:
                ok = False
                Lark(constant.LARK_CHAT_ID_P0).send_rich_text('ChatGPT reviews stop: {}'.format(excel_name),
                                                              'batch failed: {}'.format(reply))
                bot_client.send_text_with_title('user_id', req.user_id, 'ChatGPT reviews stop: {}'.format(excel_name),
                                                'batch failed: {}'.format(reply))
                break
            if not script.check_reply_ok(reply):
                Lark(constant.LARK_CHAT_ID_P0).send_rich_text('ChatGPT reviews extract stop: {}'.format(excel_name),
                                                              'batch failed for error point format: {}'.format(reply))
                bot_client.send_text_with_title('user_id', req.user_id, 'ChatGPT reviews extract stop: {}'.format(excel_name),
                                                'batch failed for error point format: {}'.format(reply))
                break

            replies.append(reply)
            if req.use_lark:
                # lark notify
                title = 'Reviews Part {} ({}: {}-{}), extract: {}'.format(i+1, excel_name, body['index_from']+1, body['index_to'], req.extract_type.name())
                msg = reply
                # if req.auto_merge:
                #     msg = 'Done'
                bot_client.send_text_with_title('user_id', req.user_id, title, msg)

        # write process info
        write_process_info(req, replies=replies)

        # 5. if auto-merge
        if req.auto_merge and ok:
            req_merge = RequestMergeReview()
            req_merge.user_id = req.user_id
            req_merge.replies = replies
            req_merge.merge_prompt = req.merge_prompt
            req_merge.system_prompt = req.system_prompt
            req_merge.temperature = req.temperature
            req_merge.use_lark = req.use_lark
            req_merge.map_reviews = map_reviews
            req_merge.archive_name = excel_name
            req_merge.extract_type = req.extract_type
            # GPT组合使用方案
            report_gpt = merge_replies_by_chatgpt(req_merge)
            # 算法Embedding模型方案
            report_algo = merge_replies_by_algo(req_merge)
            # write process info
            write_process_info(req, reports=[{'gpt': report_gpt}, {'algo': report_algo}])

    except:
        logger.error('[chatGPT run reviews] exception: {}'.format(traceback.format_exc()))
        Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT run reviews] exception', traceback.format_exc())
    finally:
        # 6. del redis lock
        redis_util.del_(redis_key.lock_chatgpt_review_run(req.user_id))

    if not ok:
        replies = None
    return replies


def merge_replies_by_algo(req: RequestMergeReview):
    if not req.use_lark or not req.replies:
        return False
    title = 'Reviews Report Algo: {}, extract: {}'.format(req.archive_name, req.extract_type.name())
    if len(req.replies) == 1:
        algo_point_list = script.parse_analyze(req.replies[0])
    else:
        # 使用算法embedding api进行合并结果
        algo_point_list = algo_client.merge_points(req.replies)
    return send_lark_result(req.user_id, title, algo_point_list)


def merge_replies_by_chatgpt(req: RequestMergeReview):
    reply, ok = None, False
    reply, cursor, loop_max, loop, filter_quote = '', 0, 100, 0, 0
    try:
        if not req.replies or not req.merge_prompt:
            return None
        excel_name = ''
        extract_type = req.extract_type.name()
        record = LarkUserWorkbench.objects.filter(lark_user_id=req.user_id, work_symbol=dto.WORK_SYMBOL).first()
        if record:
            excel_name = record.detail.get(dto.JSON_KEY_EXCEL_NAME, '')
            record.detail[dto.JSON_KEY_MERGE_PROMPT] = req.merge_prompt
            record.detail[dto.JSON_KEY_SYS_PROMPT] = req.system_prompt
            record.detail[dto.JSON_KEY_TEMPERATURE] = req.temperature
            record.save()

        replies = []
        if len(req.replies) == 1:
            # 仅有一条不用merge
            ok = True
            cursor += 1
            # 排序
            one_reply = ''
            points = script.parse_analyze(req.replies[0])
            for each in points:
                one_reply += '- {} ({})\n'.format(each['point'], ', '.join(each['ids']))
            if req.use_lark:
                bot_client.send_text_with_title('user_id', req.user_id,
                                                'Reviews Report: {}'.format(excel_name), one_reply)
            replies.append({'index_from': 1, 'index_to': 1, 'reply': one_reply})
        else:
            while cursor < len(req.replies) and loop < loop_max:
                loop += 1
                time.sleep(0.7 + random.randint(1, 5))
                ok, reply, cursor, filter_quote = each_merge(reply, cursor, req, filter_quote, loop=loop)
                if ok is False:
                    break
                else:
                    # 累积
                    replies.append({'index_from': 1, 'index_to': cursor, 'reply': reply})
                    if req.use_lark:
                        title = 'Reviews Report: {} (1-{}), extract: {}, filter_quote: {}'.format(excel_name, cursor, extract_type, filter_quote)
                        bot_client.send_text_with_title('user_id', req.user_id, title, reply)
        # 最后一次结果要trim检查，只取前20条
        trim_reply = ''
        if req.use_lark and replies:
            result = replies[-1]
            point_list = script.parse_analyze(result['reply'])
            point_list, _ = self_tag.check_report(point_list, req.map_reviews, review_meta.CHECK_TOP_COUNT, auto_trim=True)
            title = 'Reviews Report Trimmed: {} (1-{}), extract: {}, filter_quote: {}'.format(excel_name, cursor, extract_type, filter_quote)
            trim_reply = send_lark_result(req.user_id, title, point_list)

        # 再次用GPT-4合并
        if req.use_lark and trim_reply:
            req_merge_again = copy.deepcopy(req)
            req_merge_again.replies = []
            ok, reply, _, _ = each_merge(trim_reply, 0, req_merge_again, 0, model=meta.MODEL_CHAT_4)
            title = 'Reviews Report Merged Again: {} (1-{}), extract: {}, filter_quote: {}'.format(excel_name, cursor, extract_type, filter_quote)
            bot_client.send_text_with_title('user_id', req_merge_again.user_id, title, reply, at=req.user_id)

        if ok is False:
            bot_client.send_text_with_title('user_id', req.user_id, 'Reviews Report: {}'.format(excel_name), '执行失败，寻求系统管理员协助')

    except:
        logger.error('[chatGPT merge reviews] exception: {}'.format(traceback.format_exc()))
        Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT merge reviews] exception', traceback.format_exc())
    finally:
        redis_util.del_(redis_key.lock_chatgpt_review_merge(req.user_id))

    return reply


def each_merge(reply, cursor, req, filter_quote, model=meta.MODEL_CHAT_3_5, loop=None):
    # parse points and hide massive id list
    item_index, map_reviews, words, batch_content = 0, {}, 0, []
    if reply:
        # 加速merge tradeoff
        if loop and loop >= 3 and filter_quote == 0:
            filter_quote = 1
        if loop and loop >= 7 and filter_quote <= 1:
            filter_quote = 2
        text, map_reviews_each = script.transform_analyze(reply, item_index, filter_quote=filter_quote)
        item_index += len(map_reviews_each)
        map_reviews.update(map_reviews_each)
        words += len(encoding.encode(text))
        batch_content.append(text)
    for i in range(cursor, len(req.replies)):
        each = req.replies[i]
        text, map_reviews_each = script.transform_analyze(each, item_index)
        item_index += len(map_reviews_each)
        map_reviews.update(map_reviews_each)
        words += len(encoding.encode(text))
        if words <= MAX_TOKENS_IN_MERGE and len(batch_content) <= 20:
            batch_content.append(text)
            cursor += 1
        elif len(batch_content) == 1:
            # 清理提及次数少的
            return each_merge(reply, cursor, req, filter_quote=filter_quote+1)
        else:
            break
    if len(batch_content) == 1 and req.replies and cursor != len(req.replies)-1:
        logger.error('[chatGPT merge review] failed, {}'.format('single merge reply exceed'))
        Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT merge review] failed {}'.format(req.user_id),
                                                      'single merge reply exceed')
        return None, reply, cursor, filter_quote

    user_prompt = '{}\n\n{}'.format(req.merge_prompt, '\n'.join(batch_content))
    prompt_list = ['user::{}'.format(user_prompt.strip())]
    retry_max = 3
    hit, retry = False, 0
    while retry < retry_max:
        retry += 1
        hit, reply = chat_client.ask(prompt_list, system_prompt=req.system_prompt, model=model,
                                     temperature=req.temperature, auto_retry=True)
        if not script.transform_result(reply, map_reviews) and len(prompt_list) == 1:
            prompt_list.append('assistant::{}'.format(reply))
            prompt_list.append('user::{}'.format('try again, notice the format requirement.'))
            model = meta.MODEL_REVIEW
            continue
        if hit and reply and script.check_reply_ok(reply):
            break
        else:
            model = meta.MODEL_REVIEW
    # wrap reply
    if hit:
        reply = script.transform_result(reply, map_reviews)
        if not reply:
            logger.error('[chatGPT merge review] reply empty, {} at {}'.format(req.user_id, cursor))
            Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT merge review] reply empty {}'.format(req.user_id),
                                                          'cursor: {}'.format(cursor))
        return True, reply, cursor, filter_quote
    else:
        logger.error('[chatGPT merge review] failed, {}'.format(reply))
        Lark(constant.LARK_CHAT_ID_P0).send_rich_text('[chatGPT merge review] failed {}'.format(req.user_id),
                                                      reply)
        return False, None, None, filter_quote


def send_lark_result(user_id, title, point_list):
    trim_reply = ''
    for each in point_list:
        if len(each['ids']) > 0:
            trim_reply += '- {} ({})\n'.format(each['point'], ', '.join(each['ids']))
    bot_client.send_text_with_title('user_id', user_id, title, trim_reply)
    return trim_reply


if __name__ == '__main__':
    encoding = tiktoken.encoding_for_model(USE_MODEL)
    print(len(encoding.encode('hello how are you')))
