import openpyxl

from home.lark_client import bot

map_department = {}


def write_user():
    file_ = '/Users/liuqingliang/Downloads/chat_users.xlsx'
    wb = openpyxl.load_workbook(file_)
    ws = wb.active
    for i in range(2, ws.max_row+1):
        user_id = ws.cell(i, 1).value
        if not user_id:
            continue
        username = ws.cell(i, 2).value
        if username:
            continue
        user_info = bot.get_user(user_id)
        if not user_info:
            continue
        user_info = user_info['user']
        username = user_info['name']
        email = user_info.get('enterprise_email', None)
        if not email:
            email = user_info.get('email', '')
        department = ''
        department_id = user_info.get('department_ids', [''])[0]
        if department_id:
            if department_id not in map_department:
                map_department[department_id] = bot.get_department_name_path(department_id)
            department = map_department[department_id]
        print(i, username, email, department)
        ws.cell(i, 2).value = username
        ws.cell(i, 3).value = department
        ws.cell(i, 4).value = email
    wb.save(file_)


def write_user_chats():
    map_user_info = {}
    wb_user = openpyxl.load_workbook('/Users/liuqingliang/Downloads/chat_users.xlsx')
    ws_user = wb_user.active
    for i in range(2, ws_user.max_row+1):
        user_id = ws_user.cell(i, 1).value
        username = ws_user.cell(i, 2).value
        department = ws_user.cell(i, 3).value
        map_user_info[user_id] = {'name': username, 'department': department}

    obj_file = '/Users/liuqingliang/Downloads/chat_statistics_6月.xlsx'
    wb = openpyxl.load_workbook(obj_file)
    ws = wb.active
    for i in range(2, ws.max_row+1):
        user_id = ws.cell(i, 1).value
        ask_count = ws.cell(i, 2).value
        dev_count = ws.cell(i, 3).value
        print(i, user_id, ask_count, dev_count)
        user_info = map_user_info.get(user_id, {})
        ws.cell(i, 4).value = user_info['name']
        ws.cell(i, 5).value = user_info['department']
    wb.save(obj_file)


if __name__ == '__main__':
    # write_user()
    write_user_chats()

