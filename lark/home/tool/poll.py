import openpyxl
import os
import django
os.environ['DJANGO_SETTINGS_MODULE'] = 'lark.settings'
django.setup()
from home.config import constant
from home.models import LarkPoll

POLL_TYPE = 'mbti-93'
QUESTIONS = [{"index": 1, "question": "\u5f53\u4f60\u8981\u5916\u51fa\u4e00\u6574\u5929\uff0c\u4f60\u4f1a:", "option_1": "A \u8ba1\u5212\u4f60\u8981\u505a\u4ec0\u4e48\u548c\u5728\u4ec0\u4e48\u65f6\u5019\u505a", "option_2": "B \u8bf4\u53bb\u5c31\u53bb"}, {"index": 2, "question": "\u4f60\u8ba4\u4e3a\u81ea\u5df1\u662f\u4e00\u4e2a:", "option_1": "A \u8f83\u4e3a\u968f\u5174\u6240\u81f3\u7684\u4eba", "option_2": "B \u8f83\u4e3a\u6709\u6761\u7406\u7684\u4eba"}, {"index": 3, "question": "\u5047\u5982\u4f60\u662f\u4e00\u4f4d\u8001\u5e08\uff0c\u4f60\u4f1a\u9009\u6559", "option_1": "A\u4ee5\u4e8b\u5b9e\u4e3a\u4e3b\u7684\u8bfe\u7a0b", "option_2": "B\u6d89\u53ca\u7406\u8bba\u7684\u8bfe\u7a0b"}, {"index": 4, "question": "\u4f60\u901a\u5e38", "option_1": "A\u4e0e\u4eba\u5bb9\u6613\u6df7\u719f", "option_2": "B\u6bd4\u8f83\u6c89\u9759\u6216\u77dc\u6301"}, {"index": 5, "question": "\u4e00\u822c\u6765\u8bf4\uff0c\u4f60\u548c\u54ea\u4e9b\u4eba\u6bd4\u8f83\u5408\u5f97\u6765\uff1f", "option_1": "A\u5bcc\u4e8e\u60f3\u8c61\u529b\u7684\u4eba", "option_2": "B\u73b0\u5b9e\u7684\u4eba"}, {"index": 6, "question": "\u4f60\u662f\u5426\u7ecf\u5e38\u8ba9", "option_1": "A\u4f60\u7684\u60c5\u611f\u652f\u914d\u4f60\u7684\u7406\u667a", "option_2": "B\u4f60\u7684\u7406\u667a\u4e3b\u5bb0\u4f60\u7684\u60c5\u611f"}, {"index": 7, "question": "\u5904\u7406\u8bb8\u591a\u4e8b\u60c5\u4e0a\uff0c\u4f60\u4f1a\u559c\u6b22", "option_1": "A\u51ed\u5174\u6240\u81f3\u884c\u4e8b", "option_2": "B\u6309\u7167\u8ba1\u5212\u884c\u4e8b"}, {"index": 8, "question": "\u4f60\u662f\u5426", "option_1": "A\u5bb9\u6613\u8ba9\u4eba\u4e86\u89e3", "option_2": "B\u96be\u4e8e\u8ba9\u4eba\u4e86\u89e3"}, {"index": 9, "question": "\u6309\u7167\u7a0b\u5e8f\u8868\u505a\u4e8b\uff0c", "option_1": "A\u5408\u4f60\u5fc3\u610f", "option_2": "B\u4ee4\u4f60\u611f\u5230\u675f\u7f1a"}, {"index": 10, "question": "\u5f53\u4f60\u6709\u4e00\u4efd\u7279\u522b\u7684\u4efb\u52a1\uff0c\u4f60\u4f1a\u559c\u6b22", "option_1": "A\u5f00\u59cb\u524d\u5c0f\u5fc3\u7ec4\u7ec7\u8ba1\u5212", "option_2": "B\u8fb9\u505a\u8fb9\u627e\u987b\u505a\u4ec0\u4e48"}, {"index": 11, "question": "\u5728\u5927\u591a\u6570\u60c5\u51b5\u4e0b\uff0c\u4f60\u4f1a\u9009\u62e9", "option_1": "A\u987a\u5176\u81ea\u7136", "option_2": "B\u6309\u7a0b\u5e8f\u8868\u505a\u4e8b"}, {"index": 12, "question": "\u5927\u591a\u6570\u4eba\u4f1a\u8bf4\u4f60\u662f\u4e00\u4e2a", "option_1": "A\u91cd\u89c6\u81ea\u6211\u9690\u79c1\u7684\u4eba", "option_2": "B\u975e\u5e38\u5766\u7387\u5f00\u653e\u7684\u4eba"}, {"index": 13, "question": "\u4f60\u5b81\u613f\u88ab\u4eba\u8ba4\u4e3a\u662f\u4e00\u4e2a", "option_1": "A\u5b9e\u4e8b\u6c42\u662f\u7684\u4eba", "option_2": "B\u673a\u7075\u7684\u4eba"}, {"index": 14, "question": "\u5728\u4e00\u5927\u7fa4\u4eba\u5f53\u4e2d\uff0c\u901a\u5e38\u662f", "option_1": "A\u4f60\u4ecb\u7ecd\u5927\u5bb6\u8ba4\u8bc6", "option_2": "B\u522b\u4eba\u4ecb\u7ecd\u4f60"}, {"index": 15, "question": "\u4f60\u4f1a\u8ddf\u54ea\u4e9b\u4eba\u505a\u670b\u53cb\uff1f", "option_1": "A\u5e38\u63d0\u51fa\u65b0\u4e3b\u610f\u7684", "option_2": "B\u811a\u8e0f\u5b9e\u5730\u7684"}, {"index": 16, "question": "\u4f60\u503e\u5411", "option_1": "A\u91cd\u89c6\u611f\u60c5\u591a\u4e8e\u903b\u8f91", "option_2": "B\u91cd\u89c6\u903b\u8f91\u591a\u4e8e\u611f\u60c5"}, {"index": 17, "question": "\u4f60\u6bd4\u8f83\u559c\u6b22", "option_1": "A\u5750\u89c2\u4e8b\u60c5\u53d1\u5c55\u624d\u4f5c\u8ba1\u5212", "option_2": "B\u5f88\u65e9\u5c31\u4f5c\u8ba1\u5212"}, {"index": 18, "question": "\u4f60\u559c\u6b22\u82b1\u5f88\u591a\u7684\u65f6\u95f4", "option_1": "A\u4e00\u4e2a\u4eba\u72ec\u5904", "option_2": "B\u5408\u522b\u4eba\u5728\u4e00\u8d77"}, {"index": 19, "question": "\u4e0e\u5f88\u591a\u4eba\u4e00\u8d77\u4f1a", "option_1": "A\u4ee4\u4f60\u6d3b\u529b\u57f9\u589e", "option_2": "B\u5e38\u5e38\u4ee4\u4f60\u5fc3\u529b\u6194\u60b4"}, {"index": 20, "question": "\u4f60\u6bd4\u8f83\u559c\u6b22:", "option_1": "A\u5f88\u65e9\u4fbf\u628a\u7ea6\u4f1a\u3001\u793e\u4ea4\u805a\u96c6\u7b49\u4e8b\u60c5\u5b89\u6392\u59a5\u5f53", "option_2": "B\u65e0\u62d8\u65e0\u675f\uff0c\u770b\u5f53\u65f6\u6709\u4ec0\u4e48\u597d\u73a9\u5c31\u505a\u4ec0\u4e48"}, {"index": 21, "question": "\u8ba1\u5212\u4e00\u4e2a\u65c5\u7a0b\u65f6\uff0c\u4f60\u8f83\u559c\u6b22\uff1a", "option_1": "A\u5927\u90e8\u5206\u7684\u65f6\u95f4\u90fd\u662f\u8ddf\u5f53\u5929\u7684\u611f\u89c9\u884c\u4e8b", "option_2": "B\u4e8b\u5148\u77e5\u9053\u5927\u90e8\u5206\u7684\u65e5\u5b50\u4f1a\u505a\u4ec0\u4e48"}, {"index": 22, "question": "\u5728\u793e\u4ea4\u805a\u4f1a\u4e2d\uff0c\u4f60", "option_1": "A\u6709\u65f6\u611f\u5230\u90c1\u95f7", "option_2": "B\u5e38\u5e38\u4e50\u5728\u5176\u4e2d"}, {"index": 23, "question": "\u4f60\u901a\u5e38", "option_1": "A\u548c\u522b\u4eba\u5bb9\u6613\u6df7\u719f", "option_2": "B\u8d8b\u5411\u81ea\u5904\u4e00\u9685"}, {"index": 24, "question": "\u54ea\u4e9b\u4eba\u4f1a\u66f4\u5438\u5f15\u4f60\uff1f:", "option_1": "A\u4e00\u4e2a\u601d\u60f3\u654f\u6377\u53ca\u975e\u5e38\u806a\u9896\u7684\u4eba", "option_2": "B\u5b9e\u4e8b\u6c42\u662f\uff0c\u5177\u4e30\u5bcc\u5e38\u8bc6\u7684\u4eba"}, {"index": 25, "question": "\u5728\u65e5\u5e38\u5de5\u4f5c\u4e2d\uff0c\u4f60\u4f1a :", "option_1": "A\u9887\u4e3a\u559c\u6b22\u5904\u7406\u8feb\u4f7f\u4f60\u5206\u79d2\u5fc5\u4e89\u7684\u7a81\u53d1\u4e8b\u4ef6", "option_2": "B\u901a\u5e38\u9884\u5148\u8ba1\u5212\uff0c\u4ee5\u514d\u8981\u5728\u538b\u529b\u4e0b\u5de5\u4f5c"}, {"index": 26, "question": "\u4f60\u8ba4\u4e3a\u522b\u4eba\u4e00\u822c", "option_1": "A\u8981\u82b1\u5f88\u957f\u65f6\u95f4\u624d\u8ba4\u8bc6\u4f60", "option_2": "B\u7528\u5f88\u77ed\u7684\u65f6\u95f4\u4fbf\u8ba4\u8bc6\u4f60"}, {"index": 27, "question": "", "option_1": "A\u6ce8\u91cd\u9690\u79c1", "option_2": "B\u5766\u7387\u5f00\u653e"}, {"index": 28, "question": "", "option_1": "A\u9884\u5148\u5b89\u6392\u7684", "option_2": "B\u65e0\u8ba1\u5212\u7684"}, {"index": 29, "question": "", "option_1": "A\u62bd\u8c61", "option_2": "B\u5177\u4f53"}, {"index": 30, "question": "", "option_1": "A\u6e29\u67d4", "option_2": "B\u575a\u5b9a"}, {"index": 31, "question": "", "option_1": "A\u601d\u8003", "option_2": "B\u611f\u53d7"}, {"index": 32, "question": "", "option_1": "A\u4e8b\u5b9e", "option_2": "B\u610f\u5ff5"}, {"index": 33, "question": "", "option_1": "A\u51b2\u52a8", "option_2": "B\u51b3\u5b9a"}, {"index": 34, "question": "", "option_1": "A\u70ed\u8877", "option_2": "B\u6587\u9759"}, {"index": 35, "question": "", "option_1": "A\u6587\u9759", "option_2": "B\u5916\u5411"}, {"index": 36, "question": "", "option_1": "A\u6709\u7cfb\u7edf", "option_2": "B\u968f\u610f"}, {"index": 37, "question": "", "option_1": "A\u7406\u8bba", "option_2": "B\u80af\u5b9a"}, {"index": 38, "question": "", "option_1": "A\u654f\u611f", "option_2": "B\u516c\u6b63"}, {"index": 39, "question": "", "option_1": "A\u4ee4\u4eba\u4fe1\u670d", "option_2": "B\u611f\u4eba\u7684"}, {"index": 40, "question": "", "option_1": "A\u58f0\u660e", "option_2": "B\u6982\u5ff5"}, {"index": 41, "question": "", "option_1": "A\u4e0d\u53d7\u7ea6\u675f", "option_2": "B\u9884\u5148\u5b89\u6392"}, {"index": 42, "question": "", "option_1": "A\u77dc\u6301", "option_2": "B\u5065\u8c08"}, {"index": 43, "question": "", "option_1": "A\u6709\u6761\u4e0d\u7d0a", "option_2": "B\u4e0d\u62d8\u5c0f\u8282"}, {"index": 44, "question": "", "option_1": "A\u610f\u5ff5", "option_2": "B\u5b9e\u51b5"}, {"index": 45, "question": "", "option_1": "A\u540c\u60c5\u601c\u60af", "option_2": "B\u8fdc\u89c1"}, {"index": 46, "question": "", "option_1": "A\u5229\u76ca", "option_2": "B\u795d\u798f"}, {"index": 47, "question": "", "option_1": "A\u52a1\u5b9e\u7684", "option_2": "B\u7406\u8bba\u7684"}, {"index": 48, "question": "", "option_1": "A\u670b\u53cb\u4e0d\u591a", "option_2": "B\u670b\u53cb\u4f17\u591a"}, {"index": 49, "question": "", "option_1": "A\u6709\u7cfb\u7edf", "option_2": "B\u5373\u5174"}, {"index": 50, "question": "", "option_1": "A\u5bcc\u60f3\u8c61\u7684", "option_2": "B\u4ee5\u4e8b\u8bba\u4e8b"}, {"index": 51, "question": "", "option_1": "A\u4eb2\u5207\u7684", "option_2": "B\u5ba2\u89c2\u7684"}, {"index": 52, "question": "", "option_1": "A\u5ba2\u89c2\u7684", "option_2": "B\u70ed\u60c5\u7684"}, {"index": 53, "question": "", "option_1": "A\u5efa\u9020", "option_2": "B\u53d1\u660e"}, {"index": 54, "question": "", "option_1": "A\u6587\u9759", "option_2": "B\u7231\u5408\u7fa4"}, {"index": 55, "question": "", "option_1": "A\u7406\u8bba", "option_2": "B\u4e8b\u5b9e"}, {"index": 56, "question": "", "option_1": "A\u5bcc\u540c\u60c5", "option_2": "B\u5408\u903b\u8f91"}, {"index": 57, "question": "", "option_1": "A\u5177\u5206\u6790\u529b", "option_2": "B\u591a\u6101\u5584\u611f"}, {"index": 58, "question": "", "option_1": "A\u5408\u60c5\u5408\u7406", "option_2": "B\u4ee4\u4eba\u7740\u8ff7"}, {"index": 59, "question": "\u5f53\u4f60\u8981\u5728\u4e00\u4e2a\u661f\u671f\u5185\u5b8c\u6210\u4e00\u4e2a\u5927\u9879\u76ee\uff0c\u4f60\u5728\u5f00\u59cb\u7684\u65f6\u5019\u4f1a", "option_1": "A.\u628a\u8981\u505a\u7684\u4e0d\u540c\u5de5\u4f5c\u4f9d\u6b21\u5217\u51fa", "option_2": "B.\u9a6c\u4e0a\u52a8\u5de5"}, {"index": 60, "question": "\u5728\u793e\u4ea4\u573a\u5408\u4e2d\uff0c\u4f60\u7ecf\u5e38\u4f1a\u611f\u5230\uff1a", "option_1": "A.\u4e0e\u67d0\u4e9b\u4eba\u5f88\u96be\u6253\u5f00\u8bdd\u5323\u513f\u548c\u4fdd\u6301\u5bf9\u8bdd", "option_2": "B.\u4e0e\u591a\u6570\u4eba\u90fd\u80fd\u4ece\u5bb9\u5730\u957f\u8c08"}, {"index": 61, "question": "\u8981\u505a\u8bb8\u591a\u4eba\u4e5f\u505a\u7684\u4e8b\uff0c\u4f60\u6bd4\u8f83\u559c\u6b22", "option_1": "A.\u6309\u7167\u4e00\u822c\u8ba4\u53ef\u7684\u65b9\u6cd5\u53bb\u505a", "option_2": "B.\u6784\u60f3\u4e00\u4e2a\u81ea\u5df1\u7684\u60f3\u6cd5"}, {"index": 62, "question": "\u4f60\u521a\u8ba4\u8bc6\u7684\u670b\u53cb\u80fd\u5426\u8bf4\u51fa\u4f60\u7684\u5174\u8da3\uff1f", "option_1": "A.\u9a6c\u4e0a\u53ef\u4ee5", "option_2": "B.\u8981\u5f85\u4ed6\u4eec\u771f\u6b63\u4e86\u89e3\u4f60\u4e4b\u540e\u624d\u53ef\u4ee5"}, {"index": 63, "question": "\u4f60\u901a\u5e38\u8f83\u559c\u6b22\u7684\u79d1\u76ee\u662f", "option_1": "A.\u8bb2\u6388\u6982\u5ff5\u548c\u539f\u5219\u7684", "option_2": "B.\u8bb2\u6388\u4e8b\u5b9e\u548c\u6570\u636e\u7684"}, {"index": 64, "question": "\u54ea\u4e2a\u662f\u8f83\u9ad8\u7684\u8d5e\u8a89\uff0c\u6216\u79f0\u8bb8\u4e3a\uff1f", "option_1": "A.\u4e00\u8d2f\u611f\u6027\u7684\u4eba", "option_2": "B.\u4e00\u8d2f\u7406\u6027\u7684\u4eba"}, {"index": 65, "question": "\u4f60\u8ba4\u4e3a\u6309\u7167\u7a0b\u5e8f\u8868\u505a\u4e8b\uff1a", "option_1": "A.\u6709\u65f6\u662f\u9700\u8981\u7684\uff0c\u4f46\u4e00\u822c\u6765\u8bf4\u4f60\u4e0d\u5927\u559c\u6b22\u8fd9\u6837\u505a\uff0c\u6216\u662f", "option_2": "B.\u5927\u591a\u6570\u60c5\u51b5\u4e0b\u662f\u6709\u5e2e\u52a9\u800c\u4e14\u662f\u4f60\u559c\u6b22\u505a\u7684"}, {"index": 66, "question": "\u548c\u4e00\u7fa4\u4eba\u5728\u4e00\u8d77\uff0c\u4f60\u901a\u5e38\u4f1a\u9009", "option_1": "A.\u8ddf\u4f60\u5f88\u719f\u6089\u7684\u4e2a\u522b\u4eba\u8c08\u8bdd", "option_2": "B.\u53c2\u4e0e\u5927\u4f19\u7684\u8c08\u8bdd"}, {"index": 67, "question": "\u5728\u793e\u4ea4\u805a\u4f1a\u4e0a\uff0c\u4f60\u4f1a", "option_1": "A.\u662f\u8bf4\u8bdd\u5f88\u591a\u7684\u4e00\u4e2a", "option_2": "B.\u8ba9\u522b\u4eba\u591a\u8bf4\u8bdd"}, {"index": 68, "question": "\u628a\u5468\u672b\u671f\u95f4\u8981\u5b8c\u6210\u7684\u4e8b\u5217\u6210\u6e05\u5355\uff0c\u8fd9\u4e2a\u4e3b\u610f\u4f1a", "option_1": "A.\u5408\u4f60\u610f", "option_2": "B.\u4f7f\u4f60\u63d0\u4e0d\u8d77\u52b2"}, {"index": 69, "question": "\u54ea\u4e2a\u662f\u8f83\u9ad8\u7684\u8d5e\u8a89\uff0c\u6216\u79f0\u8bb8\u4e3a", "option_1": "A.\u80fd\u5e72\u7684", "option_2": "B.\u5bcc\u6709\u540c\u60c5\u5fc3"}, {"index": 70, "question": "\u4f60\u901a\u5e38\u559c\u6b22", "option_1": "A.\u4e8b\u5148\u5b89\u6392\u4f60\u7684\u793e\u4ea4\u7ea6\u4f1a", "option_2": "B.\u968f\u5174\u4e4b\u6240\u81f3\u505a\u4e8b"}, {"index": 71, "question": "\u603b\u7684\u8bf4\u6765\uff0c\u8981\u505a\u4e00\u4e2a\u5927\u578b\u4f5c\u4e1a\u65f6\uff0c\u4f60\u4f1a\u9009", "option_1": "A.\u8fb9\u505a\u8fb9\u60f3\u8be5\u505a\u4ec0\u4e48", "option_2": "B.\u9996\u5148\u628a\u5de5\u4f5c\u6309\u6b65\u7ec6\u5206"}, {"index": 72, "question": "\u4f60\u80fd\u5426\u6ed4\u6ed4\u4e0d\u7edd\u5730\u4e0e\u4eba\u804a\u5929", "option_1": "A.\u53ea\u9650\u4e8e\u8ddf\u4f60\u6709\u5171\u540c\u5174\u8da3\u7684\u4eba", "option_2": "B.\u51e0\u4e4e\u8ddf\u4efb\u4f55\u4eba\u90fd\u53ef\u4ee5"}, {"index": 73, "question": "\u4f60\u4f1a\uff1a", "option_1": "A.\u8ddf\u968f\u4e00\u4e9b\u8bc1\u660e\u6709\u6548\u7684\u65b9\u6cd5", "option_2": "B.\u5206\u6790\u8fd8\u6709\u4ec0\u4e48\u6bdb\u75c5\uff0c\u53ca\u9488\u5bf9\u5c1a\u672a\u89e3\u51b3\u7684\u96be\u9898"}, {"index": 74, "question": "\u4e3a\u4e50\u8da3\u800c\u9605\u8bfb\u65f6\uff0c\u4f60\u4f1a", "option_1": "A.\u559c\u6b22\u5947\u7279\u6216\u521b\u65b0\u7684\u8868\u8fbe\u65b9\u5f0f", "option_2": "B.\u559c\u6b22\u4f5c\u8005\u76f4\u8bdd\u76f4\u8bf4"}, {"index": 75, "question": "\u4f60\u5b81\u613f\u66ff\u54ea\u4e00\u7c7b\u4e0a\u53f8\uff08\u6216\u8005\u8001\u5e08\uff09\u5de5\u4f5c\uff1f", "option_1": "A\u5929\u6027\u6df3\u826f\uff0c\u4f46\u5e38\u5e38\u524d\u540e\u4e0d\u4e00\u7684", "option_2": "B\u8a00\u8bcd\u5c16\u9510\u4f46\u6c38\u8fdc\u5408\u4e4e\u903b\u8f91\u7684"}, {"index": 76, "question": "\u4f60\u505a\u4e8b\u591a\u6570\u662f", "option_1": "A\u6309\u5f53\u5929\u5fc3\u60c5\u53bb\u505a", "option_2": "B\u7167\u62df\u597d\u7684\u7a0b\u5e8f\u8868\u53bb\u505a"}, {"index": 77, "question": "\u4f60\u662f\u5426\uff1a", "option_1": "A\u53ef\u4ee5\u548c\u4efb\u4f55\u4eba\u6309\u9700\u6c42\u4ece\u5bb9\u5730\u4ea4\u8c08", "option_2": "B\u53ea\u662f\u5bf9\u67d0\u4e9b\u4eba\u6216\u5728\u67d0\u79cd\u60c5\u51b5\u4e0b\u624d\u53ef\u4ee5\u7545\u6240\u6b32\u8a00"}, {"index": 78, "question": "\u8981\u4f5c\u51b3\u5b9a\u65f6\uff0c\u4f60\u8ba4\u4e3a\u6bd4\u8f83\u91cd\u8981\u7684\u662f", "option_1": "A\u636e\u4e8b\u5b9e\u8861\u91cf", "option_2": "B\u8003\u8651\u4ed6\u4eba\u7684\u611f\u53d7\u548c\u610f\u89c1"}, {"index": 79, "question": "", "option_1": "A\u60f3\u8c61\u7684", "option_2": "B\u771f\u5b9e\u7684"}, {"index": 80, "question": "", "option_1": "A\u4ec1\u6148\u6177\u6168\u7684", "option_2": "B\u610f\u5fd7\u575a\u5b9a\u7684"}, {"index": 81, "question": "", "option_1": "A\u516c\u6b63\u7684", "option_2": "B\u6709\u5173\u6000\u5fc3"}, {"index": 82, "question": "", "option_1": "A\u5236\u4f5c", "option_2": "B\u8bbe\u8ba1"}, {"index": 83, "question": "", "option_1": "A\u53ef\u80fd\u6027", "option_2": "B\u5fc5\u7136\u6027"}, {"index": 84, "question": "", "option_1": "A\u6e29\u67d4", "option_2": "B\u529b\u91cf"}, {"index": 85, "question": "", "option_1": "A\u5b9e\u9645", "option_2": "B\u591a\u6101\u5584\u611f"}, {"index": 86, "question": "", "option_1": "A\u5236\u9020", "option_2": "B\u521b\u9020"}, {"index": 87, "question": "", "option_1": "A\u65b0\u9896\u7684", "option_2": "B\u5df2\u77e5\u7684"}, {"index": 88, "question": "", "option_1": "A\u540c\u60c5", "option_2": "B\u5206\u6790"}, {"index": 89, "question": "", "option_1": "A\u575a\u6301\u5df1\u89c1", "option_2": "B\u6e29\u67d4\u6709\u7231\u5fc3"}, {"index": 90, "question": "", "option_1": "A\u5177\u4f53\u7684", "option_2": "B\u62bd\u8c61\u7684"}, {"index": 91, "question": "", "option_1": "A\u5168\u5fc3\u6295\u5165", "option_2": "B\u6709\u51b3\u5fc3\u7684"}, {"index": 92, "question": "", "option_1": "A\u80fd\u5e72", "option_2": "B\u4ec1\u6148"}, {"index": 93, "question": "", "option_1": "A\u5b9e\u9645", "option_2": "B\u521b\u65b0"}]
MAP_INDEX_FACTOR = {1: ['J', 'P'], 2: ['P', 'J'], 3: ['S', 'N'], 4: ['E', 'I'], 5: ['N', 'S'], 6: ['F', 'T'], 7: ['P', 'J'], 8: ['E', 'I'], 9: ['J', 'P'], 10: ['J', 'P'], 11: ['P', 'J'], 12: ['I', 'E'], 13: ['S', 'N'], 14: ['E', 'I'], 15: ['N', 'S'], 16: ['F', 'T'], 17: ['P', 'J'], 18: ['I', 'E'], 19: ['E', 'I'], 20: ['J', 'P'], 21: ['P', 'J'], 22: ['I', 'E'], 23: ['E', 'I'], 24: ['N', 'S'], 25: ['P', 'J'], 26: ['I', 'E'], 27: ['I', 'E'], 28: ['J', 'P'], 29: ['N', 'S'], 30: ['F', 'T'], 31: ['T', 'F'], 32: ['S', 'N'], 33: ['P', 'J'], 34: ['E', 'I'], 35: ['I', 'E'], 36: ['J', 'P'], 37: ['N', 'S'], 38: ['F', 'T'], 39: ['T', 'F'], 40: ['S', 'N'], 41: ['P', 'J'], 42: ['I', 'E'], 43: ['J', 'P'], 44: ['N', 'S'], 45: ['F', 'T'], 46: ['T', 'F'], 47: ['S', 'N'], 48: ['I', 'E'], 49: ['J', 'P'], 50: ['N', 'S'], 51: ['F', 'T'], 52: ['T', 'F'], 53: ['S', 'N'], 54: ['I', 'E'], 55: ['N', 'S'], 56: ['F', 'T'], 57: ['T', 'F'], 58: ['S', 'N'], 59: ['J', 'P'], 60: ['I', 'E'], 61: ['S', 'N'], 62: ['E', 'I'], 63: ['N', 'S'], 64: ['F', 'T'], 65: ['P', 'J'], 66: ['I', 'E'], 67: ['E', 'I'], 68: ['J', 'P'], 69: ['T', 'F'], 70: ['J', 'P'], 71: ['P', 'J'], 72: ['I', 'E'], 73: ['S', 'N'], 74: ['N', 'S'], 75: ['T', 'N'], 76: ['P', 'J'], 77: ['E', 'I'], 78: ['T', 'F'], 79: ['N', 'S'], 80: ['F', 'T'], 81: ['T', 'F'], 82: ['S', 'N'], 83: ['N', 'S'], 84: ['F', 'T'], 85: ['T', 'F'], 86: ['S', 'N'], 87: ['N', 'S'], 88: ['F', 'T'], 89: ['T', 'F'], 90: ['S', 'N'], 91: ['F', 'T'], 92: ['T', 'F'], 93: ['S', 'N']}
PAIRS = [['E', 'I'], ['S', 'N'], ['T', 'F'], ['J', 'P']]
MAP_FACTOR = {'E': '外向', 'I': '内向', 'S': '感觉', 'N': '直觉', 'T': '思考', 'F': '情感', 'J': '判断', 'P': '感知'}
MAP_MBTI = {'ISTJ': '1.严肃、安静、藉由集中心 志与全力投入、及可被信赖获致成功。\n2.行事务实、有序、实际 、 逻辑、真实及可信赖\n3.十分留意且乐于任何事（工作、居家、生活均有良好组织及有序。\n4.负责任。\n5.照设定成效来作出决策且不畏阻挠与闲言会坚定为之。\n6.重视传统与忠诚。\n7.传统性的思考者或经理。', 'ISFJ': '1.安静、和善、负责任且有良心。\n2.行事尽责投入。\n3.安定性高，常居项目工作或团体之安定力量。\n4.愿投入、吃苦及力求精确。\n5.兴趣通常不在于科技方面。对细节事务有耐心。\n6.忠诚、考虑周到、知性且会关切他人感受。\n7.致力于创构有序及和谐的工作与家庭环境。', 'INFJ': '1.因为坚忍、创意及必须达成的意图而能成功。\n2.会在工作中投注最大的努力。\n3.默默强力的、诚挚的及用心的关切他人。\n4.因坚守原则而受敬重。\n5.提出造福大众利益的明确远景而为人所尊敬与追随。\n6.追求创见、关系及物质财物的意义及关联。\n7.想了解什么能激励别人及对他人具洞察力。\n8.光明正大且坚信其价值观。\n9.有组织且果断地履行其愿景。', 'INTJ': '1.具强大动力与本意来达成目的与创意—固执顽固者。\n2.有宏大的愿景且能快速在众多外界事件中找出有意义的模范。\n3.对所承负职务，具良好能力于策划工作并完成。\n4.具怀疑心、挑剔性、独立性、果决，对专业水准及绩效要求高。', 'ISTP': '1.冷静旁观者—安静、预留余地、弹性及会以无偏见的好奇心与未预期原始的幽默观察与分析。\n2.有兴趣于探索原因及效果，技术事件是为何及如何运作且使用逻辑的原理组构事实、重视效能。\n3.擅长于掌握问题核心及找出解决方式。\n4.分析成事的缘由且能实时由大量资料中找出实际问题的核心。', 'ISFP': '1.羞怯的、安宁和善地、敏感的、亲切的、且行事谦虚。\n2.喜于避开争论，不对他人强加已见或价值观。\n3.无意于领导却常是忠诚的追随者。\n4.办事不急躁，安于现状无意于以过度的急切或努力破坏现况，且非成果导向。\n5.喜欢有自有的空间及照自订的时程办事。', 'INFP': '1安静观察者，具理想性与对其价值观及重要之人具忠诚心。\n2.希外在生活形态与内在价值观相吻合。\n3.具好奇心且很快能看出机会所在。常担负开发创意的触媒者\u3000。\n4.除非价值观受侵犯，行事会具弹性、适应力高且承受力强。\n5.具想了解及发展他人潜能的企图。想作太多且作事全神贯注\u3000。\n6.对所处境遇及拥有不太在意。\n7.具适应力、有弹性除非价值观受到威胁。', 'INTP': '1.安静、自持、弹性及具适应力。\n2.特别喜爱追求理论与科学事理。\n3.习于以逻辑及分析来解决问题—问题解决者。\n4.最有兴趣于创意事务及特定工作，对聚会与闲聊无\u3000大兴趣。\n5.追求可发挥个人强烈兴趣的生涯。\n6.追求发展对有兴趣事务之逻辑解释。', 'ESTP': '1.擅长现场实时解决问题—解决问题者。\n2.喜欢办事并乐于其中及过程。\n3.倾向于喜好技术事务及运动，交结同好友人。\n4.具适应性、容忍度、务实性；投注心力于会很快具\u3000成效工作。\n5.不喜欢冗长概念的解释及理论。\n6.最专精于可操作、处理、分解或组合的真实事务。', 'ESFP': '1.外向、和善、接受性、乐于分享喜乐予他人。\n2.喜欢与他人一起行动且促成事件发生，在学习时亦然。\n3.知晓事件未来的发展并会热列参与。\n5.最擅长于人际相处能力及具备完备常识，很有弹性能立即\u3000适应他人与环境。\n6.对生命、人、物质享受的热爱者。', 'ENFP': '1.充满热忱、活力充沛、聪明的、富想象力的，视生命充满机会但期能得自他人肯定与支持。\n2.几乎能达成所有有兴趣的事。\n3.对难题很快就有对策并能对有困难的人施予援手。\n4.依赖能改善的能力而无须预作规划准备。\n5.为达目的常能找出强制自己为之的理由。\n6.即兴执行者。', 'ENTP': '1.反应快、聪明、长于多样事务。\n2.具激励伙伴、敏捷及直言讳专长。\n3.会为了有趣对问题的两面加予争辩。\n4.对解决新及挑战性的问题富有策略，但会轻忽或厌烦经常的任务与细节。\n5.兴趣多元，易倾向于转移至新生的兴趣。\n6.对所想要的会有技巧地找出逻辑的理由。\n7.长于看清础他人，有智能去解决新或有挑战的问题', 'ESTJ': '1.务实、真实、事实倾向，具企业或技术天份。\n2.不喜欢抽象理论；最喜欢学习可立即运用事理。\n3.喜好组织与管理活动且专注以最有效率方式行事以达致成效。\n4.具决断力、关注细节且很快作出决策—优秀行政者。\n5.会忽略他人感受。\n6.喜作领导者或企业主管。', 'ESFJ': '1.诚挚、爱说话、合作性高、受\u3000欢迎、光明正大 的—天生的\u3000合作者及活跃的组织成员。\n2.重和谐且长于创造和谐。\n3.常作对他人有益事务。\n4.给予鼓励及称许会有更佳工作成效。\n5.最有兴趣于会直接及有形影响人们生活的事务。\n6.喜欢与他人共事去精确且准时地完成工作。', 'ENFJ': '1.热忱、易感应及负责任的--具能鼓励他人的领导风格。\n2.对别人所想或希求会表达真正关切且切实用心去处理。\n3.能怡然且技巧性地带领团体讨论或演示文稿提案。\n4.爱交际、受欢迎及富同情心。\n5.对称许及批评很在意。\n6.喜欢带引别人且能使别人或团体发挥潜能。', 'ENTJ': '1.坦诚、具决策力的活动领导者。\n2.长于发展与实施广泛的系统以解决组织的问题。\n3.专精于具内涵与智能的谈话如对公众演讲。\n4.乐于经常吸收新知且能广开信息管道。\n5.易生过度自信，会强于表达自已创见。\n6.喜于长程策划及目标设定'}


def read_mbti_excel():
    wb = openpyxl.load_workbook('/Users/liuqingliang/Downloads/MBTI职业性格测试(自动计算版).xlsx')
    ws = wb.active
    items = []
    count = 0
    wait_option = False
    for i in range(3, 188):
        item = ws.cell(i, 2).value
        if not item:
            continue
        item = item.strip()
        if 'B' not in item:
            count += 1
            wait_option = True
            option_1 = ''
            if 'A' in item:
                tmp = item.split('A')
                item = tmp[0].strip(',， \n').strip()
                option_1 = 'A' + tmp[-1].strip(',， \n').strip()
            items.append({'index': count, 'question': item, 'option_1': option_1})
        else:
            options = item.split('B')
            option_1 = options[0].strip(',， \n').strip()
            option_2 = 'B' + options[1] if len(options) == 2 else None
            if not wait_option:
                count += 1
                items.append({'index': count, 'question': ''})
            if option_1:
                items[-1]['option_1'] = option_1
            if option_2:
                items[-1]['option_2'] = option_2
            wait_option = False
    return items


def read_mbti_formula():
    map_index_factor = {}
    wb = openpyxl.load_workbook('/Users/liuqingliang/Downloads/MBTI职业性格测试(自动计算版).xlsx')
    ws = wb['统计']
    for i in range(2, 188):
        val_b = ws.cell(i, 2).value
        val_c = ws.cell(i, 3).value
        val_d = ws.cell(i, 4).value
        val_e = ws.cell(i, 5).value
        val_f = ws.cell(i, 6).value
        val_g = ws.cell(i, 7).value
        val_h = ws.cell(i, 8).value
        val_i = ws.cell(i, 9).value
        index = int(i / 2)
        if i % 2 == 0:
            print(index)
            map_index_factor[index] = []
        factor = None
        if val_b is not None:
            print('B')
            factor = 'E'
        if val_c is not None:
            print('C')
            factor = 'I'
        if val_d is not None:
            print('D')
            factor = 'S'
        if val_e is not None:
            print('E')
            factor = 'N'
        if val_f is not None:
            print('F')
            factor = 'T'
        if val_g is not None:
            print('G')
            factor = 'F'
        if val_h is not None:
            print('H')
            factor = 'J'
        if val_i is not None:
            print('I')
            factor = 'P'
        map_index_factor[index].append(factor)

    print(map_index_factor)
    return map_index_factor


def read_mbti_type():
    map_result = {}
    wb = openpyxl.load_workbook('/Users/liuqingliang/Downloads/MBTI职业性格测试(自动计算版).xlsx')
    ws = wb['统计']
    for i in range(2, 18):
        type_ = ws.cell(i, 20).value
        desc = ws.cell(i, 21).value
        map_result[type_.strip()] = desc.strip()
    print(map_result)


def calculate_mbti(login_id, options):
    # sum
    map_result = {'E': 0, 'I': 0, 'S': 0, 'N': 0, 'T': 0, 'F': 0, 'J': 0, 'P': 0}
    if not options:
        return False, 'empty submit'
    if len(options) < len(QUESTIONS):
        return False, 'not enough submit'
    for i in range(len(options)):
        index = i + 1
        if options[i] == 'A':
            value = 0
        elif options[i] == 'B':
            value = 1
        else:
            return False, 'error submit'
        obj_factor = MAP_INDEX_FACTOR[index][value]
        map_result[obj_factor] += 1
    # cal
    print(map_result)
    mbti = ''
    for pair in PAIRS:
        if map_result[pair[0]] > map_result[pair[1]]:
            mbti += pair[0]
        else:
            mbti += pair[1]
    print(mbti)
    print(MAP_MBTI[mbti])
    # save
    record = LarkPoll.objects.filter(poll_type=POLL_TYPE, lark_user=login_id).first()
    if record is None:
        record = LarkPoll(poll_type=POLL_TYPE, lark_user=login_id)
    record.repoll = constant.NO_IN_DB
    record.answer = {'answer': options}
    record.result = {'mbti': mbti, 'score': map_result, 'desc': MAP_MBTI[mbti]}
    record.save()
    return True, ''


if __name__ == '__main__':
    calculate_mbti('', ['A', 'B', 'B', 'B', 'A', 'B', 'B', 'B', 'B', 'A', 'A', 'B', 'B', 'B', 'B', 'B', 'B', 'A', 'B', 'A', 'B', 'A', 'B', 'A', 'B', 'A', 'B', 'A', 'A', 'A', 'A', 'B', 'B', 'B', 'A', 'A', 'A', 'B', 'A', 'B', 'B', 'A', 'A', 'A', 'B', 'A', 'A', 'A', 'A', 'A', 'B', 'A', 'A', 'A', 'A', 'B', 'A', 'A', 'A', 'A', 'B', 'B', 'A', 'B', 'A', 'A', 'B', 'A', 'A', 'A', 'B', 'A', 'B', 'B', 'B', 'B', 'B', 'A', 'B', 'B', 'A', 'B', 'B', 'B', 'A', 'B', 'A', 'B', 'B', 'B', 'B', 'A', 'B'])

